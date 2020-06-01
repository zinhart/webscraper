import platform
import winreg
import sys
import os
from time import sleep
from selenium import webdriver
from openpyxl import Workbook, load_workbook
import datetime
import defs


def get_os():
    return str.lower(platform.system())
def get_all_browsers(operating_system):
    if operating_system == 'windows':
        apps = 'SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths'
        windows_browsers = defs.supported_browsers['windows']
        access_registry = winreg.ConnectRegistry(None,winreg.HKEY_LOCAL_MACHINE)
        access_key = winreg.OpenKey(access_registry, apps)
        acess_key_count = winreg.QueryInfoKey(access_key)[0]
        exe= []
        #accessing the key to open the registry directories under
        for n in range(acess_key_count):
            try:
                browser = str.lower(winreg.EnumKey(access_key, n))
                if browser in windows_browsers.keys():
                    app_path = apps+str("\\")+browser
                    handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, app_path)
                    num_values = winreg.QueryInfoKey(handle)[1]
                    for i in range(num_values):
                        paths = winreg.EnumValue(handle, i)
                        windows_browsers[browser] = paths[1]

            except:
                break
    return windows_browsers

def get_browser_path(os_name, browser_name):
    all_browsers = get_all_browsers(os_name)
    if browser_name in all_browsers.keys():
            return str(all_browsers[browser_name] + "\\" + browser_name)
    else:
        print("error unsupported browser")
        return None
def get_web_driver(os_name, browser_name):
    if os_name == 'windows':
        if browser_name == 'chrome.exe':
            defs.supported_webdrivers[os_name][browser_name] = webdriver.Chrome()
        elif browser_name == 'iexplore.exe':
            defs.supported_webdrivers[os_name][browser_name] = webdriver.Ie()
        elif browser_name == 'microsoftedge.exe':
            defs.supported_webdrivers[os_name][browser_name] = webdriver.Edge()
        return defs.supported_webdrivers[os_name][browser_name]

def login(browser_name, url, usr, pw):
    os_name = get_os()
    if os_name not in defs.supported_os:
            print("unsupported os")
            return False
    browser_path = get_browser_path(os_name, browser_name)
    if browser_path == None:
        print("browser: ",browser," not found")
        return False
    driver = get_web_driver(os_name, browser_name)
    driver.get(url)
    username = driver.find_element_by_id("username")
    password = driver.find_element_by_id("password")
    username.send_keys(usr)
    password.send_keys(pw)
    sleep(0.75)
    # xpath from chropath
    sign_in_button = driver.find_element_by_xpath('//button[@type="submit"]')
    sign_in_button.click()
    defs.webdriver = driver

    return True

def go_to_my_jobs():
    if defs.logged_in is True:
        print("going to my jobs")
        defs.webdriver.find_element_by_xpath('//a[@href="/jobs/"]').click()

def get_saved_jobs():
    if defs.logged_in is True:
        print("getting saved jobs")
        defs.webdriver.get("https://www.linkedin.com/jobs/tracker/saved/")
        rows = {'positions': [None],
                'companies': [None],
                'job locations': [None],
                'days old/current # of applicants': [None],
                'applied': [None],
                'date application sent': [None],
                'result': [None],
                'date received': [None]
               }
        scroll_pause_time = 2.0
        # Get scroll height
        last_height = defs.webdriver.execute_script("return document.body.scrollHeight")
        while True:
            position_titles = defs.webdriver.find_elements_by_xpath("//*[starts-with(@id, 'ember') and @class='jobs-job-card-content__title ember-view'] ")
            companies_names = defs.webdriver.find_elements_by_xpath("//*[starts-with(@id, 'ember') and @class='t-black jobs-job-card-content__company-name t-14 t-normal ember-view']")
            job_location = defs.webdriver.find_elements_by_xpath("//*[@class='t-12 t-black--light']")
            bullets = defs.webdriver.find_elements_by_xpath("//*[@class='jobs-job-card-content__info t-12 t-black--light mt2 pt3']")
            for (pt, cn, jl, bl) in zip(position_titles, companies_names, job_location, bullets):
                rows['positions'].append(pt.text)
                rows['companies'].append(cn.text)
                rows['job locations'].append(jl.text)
                rows['days old/current # of applicants'].append(bl.text)
                rows['applied'].append('N')
                rows['date application sent'].append("")
                rows['result'].append("")
                rows['date received'].append("")
            # Scroll down to bottom
            defs.webdriver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            # Wait to load page
            sleep(scroll_pause_time)
            # Calculate new scroll height and compare with last scroll height
            new_height = defs.webdriver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
        return rows

def get_applied_jobs():
        defs.webdriver.get("https://www.linkedin.com/jobs/tracker/applied/")
        rows = []
        rows.append(['positions', 'companies', 'job locations'])
        scroll_pause_time = 2.0
        # Get scroll height
        last_height = defs.webdriver.execute_script("return document.body.scrollHeight")
        while True:
            position_titles = defs.webdriver.find_elements_by_xpath("//*[starts-with(@id, 'ember') and @class='jobs-job-card-content__title ember-view'] ")
            companies_names = defs.webdriver.find_elements_by_xpath("//*[starts-with(@id, 'ember') and @class='t-black jobs-job-card-content__company-name t-14 t-normal ember-view']")
            job_location = defs.webdriver.find_elements_by_xpath("//*[@class='t-12 t-black--light']")
            for (pt, cn, jl) in zip(position_titles, companies_names, job_location):
                rows.append([pt.text, cn.text, jl.text])
            # Scroll down to bottom
            defs.webdriver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            # Wait to load page
            sleep(scroll_pause_time)
            # Calculate new scroll height and compare with last scroll height
            new_height = defs.webdriver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
        return rows



def generate_final_report(saved_jobs, applied_jobs):
    print("in generate final report")
    for i in  range(1, len(applied_jobs)):
        for j in range(1, len(saved_jobs)):
            if applied_jobs[i][0] == saved_jobs[j][0] and applied_jobs[i][1] == saved_jobs[j][1] and applied_jobs[i][2] == saved_jobs[j][2]:
                pass
                #print(applied_jobs[i][0], saved_jobs[j][0])
                #print(applied_jobs[i][1], saved_jobs[j][1])
                #print(applied_jobs[i][2], saved_jobs[j][2])
                #saved_jobs[j][4] = "Y"
                #print(saved_jobs[j][0]+" + " +saved_jobs[j][1] +" + " +saved_jobs[j][2])

    return None

def as_text(value):
    if value is None:
        return ""
    return str(value)
def create_workbook(filename):
    if os.path.exists(filename):
        defs.workbook = load_workbook(filename)
    else:
        defs.workbook = Workbook()
def create_sheet(sheet_name):
    defs.workbook.create_sheet(sheet_name)
def insert_column(sheet_name, column, column_name, data):
    column_label_index = column + str(1)
    defs.workbook[sheet_name][column_label_index] = column_name
    for i in range(2, len(data)):
        index = column + str(i)
        defs.workbook[sheet_name][index] = data[i]
def autofit_cells(sheet_name):#  autofit columns to data size
    for column_cells in defs.workbook[sheet_name].columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        defs.workbook[sheet_name].column_dimensions[column_cells[0].column_letter].width = length
def save(filename):
        defs.workbook.save(filename)
def main():
    if len(sys.argv) != 5:
        print("error")
        exit()
    browser_name = sys.argv[1]
    url = sys.argv[2]
    usr = sys.argv[3]
    pw  = sys.argv[4]
    print(browser_name, url, usr, pw)
    defs.logged_in  = login(browser_name, url, usr, pw)
    go_to_my_jobs()
    saved_jobs = get_saved_jobs()
    #applied_jobs = get_applied_jobs();
    #report = generate_final_report(saved_jobs, applied_jobs)
    filename = "job_search_data.xlsx"
    sheetname = datetime.datetime.now().strftime("%Y-%m-%d %H,%M,%S")
    create_workbook(filename)
    create_sheet(sheetname)
    insert_column(sheetname,"A", 'positions', saved_jobs['positions'])
    insert_column(sheetname,"B", 'companies', saved_jobs['companies'])
    insert_column(sheetname,"C", 'job locations', saved_jobs['job locations'])
    insert_column(sheetname,"D", 'days old/current # of applicants', saved_jobs['days old/current # of applicants'])
    insert_column(sheetname,"E", 'date application sent', saved_jobs['date application sent'])
    insert_column(sheetname,"F", 'result', saved_jobs['result'])
    insert_column(sheetname,"G", 'date received', saved_jobs['date received'])
    autofit_cells(sheetname)
    save(filename)

if __name__ == '__main__':
    main()
